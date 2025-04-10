"""
Rutas para la gestión de turnos y horarios de empleados.

Este módulo proporciona las rutas para:
- Crear y editar turnos
- Asignar turnos a empleados
- Gestionar ausencias
- Crear plantillas de horarios
- Visualizar horarios semanales
- Generar informes
"""

from flask import (Blueprint, render_template, redirect, url_for, flash, 
                  request, jsonify, current_app, send_file, abort, session)
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash
from app import db
from datetime import datetime, timedelta, date
import calendar
from models import Company, Employee, User, Activity, Role
from models_turnos import (Turno, Horario, Ausencia, RequisitoPersonal, PlantillaHorario, 
                         DetallePlantilla, AsignacionPlantilla, PreferenciaDisponibilidad,
                         HistorialCambios, TipoTurno, TipoAusencia)
from forms_turnos import (TurnoForm, HorarioForm, HorarioMasivoForm, EliminarHorarioForm, AusenciaForm, 
                        AprobarAusenciaForm, RequisitoPersonalForm, PlantillaHorarioForm, 
                        DetallePlantillaForm, AsignacionPlantillaForm, PreferenciaDisponibilidadForm,
                        FiltroHorarioForm)
import json
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import or_, and_, func, extract

# Crear el blueprint
turnos_bp = Blueprint('turnos', __name__, url_prefix='/turnos')

# Variables de ayuda
dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

# Funciones auxiliares
def puede_gestionar_empresa(company_id):
    """Verifica si el usuario actual puede gestionar la empresa especificada."""
    if current_user.is_admin():
        return True
    
    if current_user.is_gerente():
        for company in current_user.companies:
            if company.id == company_id:
                return True
    
    return False

def registrar_cambio_horario(horario, tipo_cambio, descripcion):
    """Registra un cambio en el historial de cambios de horarios."""
    cambio = HistorialCambios(
        tipo_cambio=tipo_cambio,
        descripcion=descripcion,
        horario_id=horario.id if horario else None,
        user_id=current_user.id
    )
    db.session.add(cambio)
    db.session.commit()

def get_semana_actual():
    """Obtiene la fecha de inicio (lunes) de la semana actual."""
    hoy = date.today()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    return inicio_semana

def get_empleados_empresa(company_id):
    """Obtiene los empleados activos de una empresa."""
    return Employee.query.filter_by(company_id=company_id, is_active=True).all()

def get_turnos_empresa(company_id):
    """Obtiene los turnos activos de una empresa."""
    return Turno.query.filter_by(company_id=company_id, is_active=True).all()

def get_plantillas_empresa(company_id):
    """Obtiene las plantillas de horario activas de una empresa."""
    return PlantillaHorario.query.filter_by(company_id=company_id, is_active=True).all()

# Rutas para la selección de empresa
@turnos_bp.route('/')
@login_required
def index():
    """Página principal del módulo de turnos."""
    if current_user.is_admin():
        companies = Company.query.filter_by(is_active=True).all()
    else:
        companies = current_user.companies
    
    return render_template('turnos/index.html', 
                          title='Módulo de Turnos', 
                          companies=companies)

@turnos_bp.route('/select_company', methods=['GET', 'POST'])
@login_required
def select_company():
    """Selección de empresa para el módulo de turnos."""
    if request.method == 'POST':
        company_id = request.form.get('company_id', type=int)
        
        if not company_id:
            flash('Debe seleccionar una empresa.', 'danger')
            return redirect(url_for('turnos.select_company'))
        
        if not puede_gestionar_empresa(company_id):
            flash('No tiene permiso para gestionar esta empresa.', 'danger')
            return redirect(url_for('turnos.select_company'))
        
        # Guardar la empresa seleccionada en la sesión
        session['turnos_company_id'] = company_id
        company = Company.query.get_or_404(company_id)
        
        return redirect(url_for('turnos.dashboard', company_id=company_id))
    
    if current_user.is_admin():
        companies = Company.query.filter_by(is_active=True).all()
    else:
        companies = current_user.companies
    
    return render_template('turnos/select_company.html', 
                          title='Seleccionar Empresa', 
                          companies=companies)

# Dashboard y visualización de horarios
@turnos_bp.route('/dashboard/<int:company_id>')
@login_required
def dashboard(company_id):
    """Dashboard principal del módulo de turnos."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    # Guardar la empresa seleccionada en la sesión
    session['turnos_company_id'] = company_id
    company = Company.query.get_or_404(company_id)
    
    # Obtener estadísticas para el dashboard
    empleados = get_empleados_empresa(company_id)
    turnos = get_turnos_empresa(company_id)
    inicio_semana = get_semana_actual()
    fin_semana = inicio_semana + timedelta(days=6)
    
    # Contar horarios asignados para la semana actual
    asignaciones_semana = Horario.query.join(Employee).filter(
        Employee.company_id == company_id,
        Horario.fecha >= inicio_semana,
        Horario.fecha <= fin_semana,
        Horario.is_active == True
    ).count()
    
    # Contar ausencias activas
    ausencias_activas = Ausencia.query.join(Employee).filter(
        Employee.company_id == company_id,
        Ausencia.fecha_fin >= date.today(),
        Ausencia.aprobado == True
    ).count()
    
    # Últimos cambios en horarios
    ultimos_cambios = HistorialCambios.query.join(
        Horario, HistorialCambios.horario_id == Horario.id
    ).join(
        Employee, Horario.employee_id == Employee.id
    ).filter(
        Employee.company_id == company_id
    ).order_by(HistorialCambios.fecha_cambio.desc()).limit(5).all()
    
    return render_template('turnos/dashboard.html',
                          title=f'Dashboard de Turnos - {company.name}',
                          company=company,
                          empleados_count=len(empleados),
                          turnos_count=len(turnos),
                          asignaciones_semana=asignaciones_semana,
                          ausencias_activas=ausencias_activas,
                          inicio_semana=inicio_semana,
                          fin_semana=fin_semana,
                          ultimos_cambios=ultimos_cambios)

@turnos_bp.route('/calendario/<int:company_id>')
@login_required
def calendario(company_id):
    """Vista de calendario semanal de turnos."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    
    # Obtener parámetros de fecha
    semana_str = request.args.get('semana')
    if semana_str:
        try:
            inicio_semana = datetime.strptime(semana_str, '%Y-%m-%d').date()
        except ValueError:
            inicio_semana = get_semana_actual()
    else:
        inicio_semana = get_semana_actual()
    
    fin_semana = inicio_semana + timedelta(days=6)
    semana_anterior = inicio_semana - timedelta(days=7)
    semana_siguiente = inicio_semana + timedelta(days=7)
    
    # Obtener empleados y turnos
    empleados = get_empleados_empresa(company_id)
    turnos = get_turnos_empresa(company_id)
    
    # Obtener horarios para la semana seleccionada
    horarios = Horario.query.join(Employee).filter(
        Employee.company_id == company_id,
        Horario.fecha >= inicio_semana,
        Horario.fecha <= fin_semana,
        Horario.is_active == True
    ).all()
    
    # Organizar horarios por empleado y día
    datos_calendario = {}
    for empleado in empleados:
        datos_calendario[empleado.id] = {
            'empleado': empleado,
            'dias': {(inicio_semana + timedelta(days=i)): [] for i in range(7)}
        }
    
    for horario in horarios:
        if horario.employee_id in datos_calendario:
            datos_calendario[horario.employee_id]['dias'][horario.fecha].append(horario)
    
    # Obtener ausencias para la semana seleccionada
    ausencias = Ausencia.query.join(Employee).filter(
        Employee.company_id == company_id,
        Ausencia.fecha_inicio <= fin_semana,
        Ausencia.fecha_fin >= inicio_semana,
        Ausencia.aprobado == True
    ).all()
    
    # Organizar ausencias por empleado y día
    datos_ausencias = {}
    for ausencia in ausencias:
        if ausencia.employee_id not in datos_ausencias:
            datos_ausencias[ausencia.employee_id] = []
        datos_ausencias[ausencia.employee_id].append(ausencia)
    
    # Calcular los días de la semana
    dias_semana_fechas = [(inicio_semana + timedelta(days=i)) for i in range(7)]
    
    # Formulario para eliminar horario
    eliminar_form = EliminarHorarioForm()
    
    return render_template('turnos/calendario.html',
                          title=f'Calendario de Turnos - {company.name}',
                          company=company,
                          inicio_semana=inicio_semana,
                          fin_semana=fin_semana,
                          semana_anterior=semana_anterior,
                          semana_siguiente=semana_siguiente,
                          dias_semana=dias_semana,
                          dias_semana_fechas=dias_semana_fechas,
                          datos_calendario=datos_calendario,
                          datos_ausencias=datos_ausencias,
                          empleados=empleados,
                          turnos=turnos,
                          eliminar_form=eliminar_form,
                          meses=meses)

# Gestión de turnos
@turnos_bp.route('/turnos/<int:company_id>')
@login_required
def listar_turnos(company_id):
    """Lista los turnos de una empresa."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    turnos = Turno.query.filter_by(company_id=company_id, is_active=True).order_by(Turno.nombre).all()
    
    return render_template('turnos/listar_turnos.html',
                          title=f'Turnos - {company.name}',
                          company=company,
                          turnos=turnos)

@turnos_bp.route('/turnos/<int:company_id>/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_turno(company_id):
    """Crea un nuevo turno."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = TurnoForm()
    
    if form.validate_on_submit():
        turno = Turno(
            nombre=form.nombre.data,
            tipo=TipoTurno[form.tipo.data],
            hora_inicio=form.hora_inicio.data,
            hora_fin=form.hora_fin.data,
            color=form.color.data,
            descripcion=form.descripcion.data,
            descanso_inicio=form.descanso_inicio.data,
            descanso_fin=form.descanso_fin.data,
            company_id=company_id
        )
        
        db.session.add(turno)
        db.session.commit()
        
        flash(f'Turno "{turno.nombre}" creado correctamente.', 'success')
        return redirect(url_for('turnos.listar_turnos', company_id=company_id))
    
    form.company_id.data = company_id
    
    return render_template('turnos/form_turno.html',
                          title=f'Nuevo Turno - {company.name}',
                          company=company,
                          form=form)

@turnos_bp.route('/turnos/<int:company_id>/editar/<int:turno_id>', methods=['GET', 'POST'])
@login_required
def editar_turno(company_id, turno_id):
    """Edita un turno existente."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    turno = Turno.query.get_or_404(turno_id)
    
    if turno.company_id != company_id:
        flash('El turno no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_turnos', company_id=company_id))
    
    form = TurnoForm(obj=turno)
    
    if form.validate_on_submit():
        form.populate_obj(turno)
        turno.tipo = TipoTurno[form.tipo.data]
        
        db.session.commit()
        
        flash(f'Turno "{turno.nombre}" actualizado correctamente.', 'success')
        return redirect(url_for('turnos.listar_turnos', company_id=company_id))
    
    if request.method == 'GET':
        form.tipo.data = turno.tipo.name
    
    return render_template('turnos/form_turno.html',
                          title=f'Editar Turno - {company.name}',
                          company=company,
                          form=form,
                          turno=turno)

@turnos_bp.route('/turnos/<int:company_id>/eliminar/<int:turno_id>', methods=['POST'])
@login_required
def eliminar_turno(company_id, turno_id):
    """Elimina un turno (marcar como inactivo)."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    turno = Turno.query.get_or_404(turno_id)
    
    if turno.company_id != company_id:
        flash('El turno no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_turnos', company_id=company_id))
    
    # Verificar si hay horarios asignados a este turno
    horarios_count = Horario.query.filter_by(turno_id=turno.id, is_active=True).count()
    if horarios_count > 0:
        flash(f'No se puede eliminar el turno "{turno.nombre}" porque tiene {horarios_count} horarios asignados.', 'danger')
        return redirect(url_for('turnos.listar_turnos', company_id=company_id))
    
    turno.is_active = False
    db.session.commit()
    
    flash(f'Turno "{turno.nombre}" eliminado correctamente.', 'success')
    return redirect(url_for('turnos.listar_turnos', company_id=company_id))

# Gestión de horarios
@turnos_bp.route('/horarios/<int:company_id>/asignar', methods=['GET', 'POST'])
@login_required
def asignar_horario(company_id):
    """Asigna un turno a un empleado."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = HorarioForm()
    
    # Cargar empleados y turnos para los select
    empleados = get_empleados_empresa(company_id)
    turnos = get_turnos_empresa(company_id)
    
    form.employee_id.choices = [(e.id, f"{e.nombre} {e.apellidos}") for e in empleados]
    form.turno_id.choices = [(t.id, t.nombre) for t in turnos]
    
    if form.validate_on_submit():
        # Verificar si ya existe una asignación para este empleado en esta fecha
        existe = Horario.query.filter_by(
            employee_id=form.employee_id.data,
            fecha=form.fecha.data,
            is_active=True
        ).first()
        
        if existe:
            flash(f'El empleado ya tiene un turno asignado para esta fecha. Por favor, elimine el turno existente primero.', 'warning')
            return redirect(url_for('turnos.asignar_horario', company_id=company_id))
        
        # Verificar si el empleado tiene ausencia en esta fecha
        empleado = Employee.query.get(form.employee_id.data)
        ausencia = Ausencia.query.filter(
            Ausencia.employee_id == form.employee_id.data,
            Ausencia.fecha_inicio <= form.fecha.data,
            Ausencia.fecha_fin >= form.fecha.data,
            Ausencia.aprobado == True
        ).first()
        
        if ausencia:
            flash(f'El empleado tiene una ausencia registrada para esta fecha ({ausencia.tipo.value}). ¿Desea asignar el turno de todas formas?', 'warning')
            # Aquí podrías agregar una confirmación adicional o simplemente permitir la asignación
        
        horario = Horario(
            fecha=form.fecha.data,
            turno_id=form.turno_id.data,
            employee_id=form.employee_id.data,
            notas=form.notas.data
        )
        
        db.session.add(horario)
        db.session.commit()
        
        # Registrar el cambio en el historial
        turno = Turno.query.get(form.turno_id.data)
        registrar_cambio_horario(
            horario,
            "creacion",
            f"Asignación de turno '{turno.nombre}' a {empleado.nombre} {empleado.apellidos} para el {form.fecha.data.strftime('%d/%m/%Y')}"
        )
        
        flash(f'Turno asignado correctamente a {empleado.nombre} {empleado.apellidos}.', 'success')
        
        # Redirigir al calendario en la semana de la fecha asignada
        inicio_semana = form.fecha.data - timedelta(days=form.fecha.data.weekday())
        return redirect(url_for('turnos.calendario', company_id=company_id, semana=inicio_semana.strftime('%Y-%m-%d')))
    
    # Si viene de un día específico del calendario, pre-cargar la fecha
    fecha_str = request.args.get('fecha')
    if fecha_str:
        try:
            form.fecha.data = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            form.fecha.data = date.today()
    else:
        form.fecha.data = date.today()
    
    # Si viene con un empleado preseleccionado
    employee_id = request.args.get('employee_id', type=int)
    if employee_id:
        form.employee_id.data = employee_id
    
    return render_template('turnos/form_horario.html',
                          title=f'Asignar Turno - {company.name}',
                          company=company,
                          form=form)

@turnos_bp.route('/horarios/<int:company_id>/asignar_masivo', methods=['GET', 'POST'])
@login_required
def asignar_horario_masivo(company_id):
    """Asigna turnos de forma masiva a varios empleados en un rango de fechas."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = HorarioMasivoForm()
    
    # Cargar empleados y turnos para los select
    empleados = get_empleados_empresa(company_id)
    turnos = get_turnos_empresa(company_id)
    
    form.employee_ids.choices = [(e.id, f"{e.nombre} {e.apellidos}") for e in empleados]
    form.turno_id.choices = [(t.id, t.nombre) for t in turnos]
    
    if form.validate_on_submit():
        fecha_inicio = form.fecha_inicio.data
        fecha_fin = form.fecha_fin.data
        turno_id = form.turno_id.data
        employee_ids = form.employee_ids.data
        dias_semana = form.dias_semana.data
        notas = form.notas.data
        
        # Crear horarios para cada combinación de empleado y fecha
        horarios_creados = 0
        horarios_omitidos = 0
        
        # Para cada fecha en el rango
        fecha_actual = fecha_inicio
        while fecha_actual <= fecha_fin:
            # Verificar si el día de la semana está incluido
            if fecha_actual.weekday() in dias_semana:
                # Para cada empleado seleccionado
                for employee_id in employee_ids:
                    # Verificar si ya existe una asignación
                    existe = Horario.query.filter_by(
                        employee_id=employee_id,
                        fecha=fecha_actual,
                        is_active=True
                    ).first()
                    
                    if not existe:
                        horario = Horario(
                            fecha=fecha_actual,
                            turno_id=turno_id,
                            employee_id=employee_id,
                            notas=notas
                        )
                        
                        db.session.add(horario)
                        horarios_creados += 1
                        
                        # Registrar el cambio en el historial (esto puede generar muchos registros)
                        # Por eficiencia, podríamos considerar registrar un único cambio masivo
                        empleado = Employee.query.get(employee_id)
                        turno = Turno.query.get(turno_id)
                        registrar_cambio_horario(
                            horario,
                            "creacion",
                            f"Asignación masiva de turno '{turno.nombre}' a {empleado.nombre} {empleado.apellidos} para el {fecha_actual.strftime('%d/%m/%Y')}"
                        )
                    else:
                        horarios_omitidos += 1
            
            # Pasar al siguiente día
            fecha_actual += timedelta(days=1)
        
        db.session.commit()
        
        mensaje = f'Se han creado {horarios_creados} asignaciones de horario correctamente.'
        if horarios_omitidos > 0:
            mensaje += f' Se omitieron {horarios_omitidos} asignaciones porque ya existían.'
        
        flash(mensaje, 'success')
        
        # Redirigir al calendario en la semana de la fecha inicial
        inicio_semana = fecha_inicio - timedelta(days=fecha_inicio.weekday())
        return redirect(url_for('turnos.calendario', company_id=company_id, semana=inicio_semana.strftime('%Y-%m-%d')))
    
    # Valores por defecto
    if request.method == 'GET':
        form.fecha_inicio.data = date.today()
        form.fecha_fin.data = date.today() + timedelta(days=6)  # Una semana
        form.dias_semana.data = [0, 1, 2, 3, 4]  # Lunes a viernes
    
    return render_template('turnos/form_horario_masivo.html',
                          title=f'Asignación Masiva de Turnos - {company.name}',
                          company=company,
                          form=form)

@turnos_bp.route('/horarios/<int:company_id>/eliminar', methods=['POST'])
@login_required
def eliminar_horario(company_id):
    """Elimina una asignación de horario."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    form = EliminarHorarioForm()
    
    if form.validate_on_submit():
        horario_id = form.horario_id.data
        horario = Horario.query.get_or_404(horario_id)
        
        # Verificar que el empleado pertenece a la empresa
        empleado = Employee.query.get(horario.employee_id)
        if empleado.company_id != company_id:
            flash('El horario no pertenece a un empleado de esta empresa.', 'danger')
            return redirect(url_for('turnos.calendario', company_id=company_id))
        
        # Guardar información para el historial antes de desactivar
        descripcion = f"Eliminación de turno '{horario.turno.nombre}' asignado a {empleado.nombre} {empleado.apellidos} para el {horario.fecha.strftime('%d/%m/%Y')}"
        
        # Marcar como inactivo en lugar de eliminar
        horario.is_active = False
        db.session.commit()
        
        # Registrar el cambio en el historial
        registrar_cambio_horario(horario, "eliminacion", descripcion)
        
        flash('Asignación de turno eliminada correctamente.', 'success')
        
        # Redirigir al calendario en la semana de la fecha del horario
        inicio_semana = horario.fecha - timedelta(days=horario.fecha.weekday())
        return redirect(url_for('turnos.calendario', company_id=company_id, semana=inicio_semana.strftime('%Y-%m-%d')))
    
    flash('No se pudo eliminar el horario.', 'danger')
    return redirect(url_for('turnos.calendario', company_id=company_id))

# Gestión de ausencias
@turnos_bp.route('/ausencias/<int:company_id>')
@login_required
def listar_ausencias(company_id):
    """Lista las ausencias de los empleados de una empresa."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    
    # Filtros
    estado = request.args.get('estado', 'activas')
    tipo = request.args.get('tipo')
    
    # Query base
    query = Ausencia.query.join(Employee).filter(Employee.company_id == company_id)
    
    # Aplicar filtros
    if estado == 'activas':
        query = query.filter(Ausencia.fecha_fin >= date.today())
    elif estado == 'pendientes':
        query = query.filter(Ausencia.aprobado == False)
    elif estado == 'aprobadas':
        query = query.filter(Ausencia.aprobado == True)
    
    if tipo:
        query = query.filter(Ausencia.tipo == TipoAusencia[tipo])
    
    # Ordenar por fecha de inicio (más recientes primero)
    ausencias = query.order_by(Ausencia.fecha_inicio.desc()).all()
    
    return render_template('turnos/listar_ausencias.html',
                          title=f'Ausencias - {company.name}',
                          company=company,
                          ausencias=ausencias,
                          estado=estado,
                          tipo=tipo,
                          tipos_ausencia=TipoAusencia)

@turnos_bp.route('/ausencias/<int:company_id>/nueva', methods=['GET', 'POST'])
@login_required
def nueva_ausencia(company_id):
    """Registra una nueva ausencia para un empleado."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = AusenciaForm()
    
    # Cargar empleados para el select
    empleados = get_empleados_empresa(company_id)
    form.employee_id.choices = [(e.id, f"{e.nombre} {e.apellidos}") for e in empleados]
    
    if form.validate_on_submit():
        ausencia = Ausencia(
            fecha_inicio=form.fecha_inicio.data,
            fecha_fin=form.fecha_fin.data,
            tipo=TipoAusencia[form.tipo.data],
            motivo=form.motivo.data,
            employee_id=form.employee_id.data
        )
        
        # Si el usuario es admin o gerente, aprobar automáticamente
        if current_user.is_admin() or current_user.is_gerente():
            ausencia.aprobado = True
            ausencia.aprobado_por_id = current_user.id
        
        db.session.add(ausencia)
        db.session.commit()
        
        empleado = Employee.query.get(form.employee_id.data)
        flash(f'Ausencia registrada correctamente para {empleado.nombre} {empleado.apellidos}.', 'success')
        return redirect(url_for('turnos.listar_ausencias', company_id=company_id))
    
    # Si viene con un empleado preseleccionado
    employee_id = request.args.get('employee_id', type=int)
    if employee_id:
        form.employee_id.data = employee_id
    
    # Valores por defecto
    if request.method == 'GET':
        form.fecha_inicio.data = date.today()
        form.fecha_fin.data = date.today()
    
    return render_template('turnos/form_ausencia.html',
                          title=f'Nueva Ausencia - {company.name}',
                          company=company,
                          form=form)

@turnos_bp.route('/ausencias/<int:company_id>/aprobar', methods=['POST'])
@login_required
def aprobar_ausencia(company_id):
    """Aprueba o rechaza una ausencia."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    form = AprobarAusenciaForm()
    
    if form.validate_on_submit():
        ausencia_id = form.ausencia_id.data
        ausencia = Ausencia.query.get_or_404(ausencia_id)
        
        # Verificar que el empleado pertenece a la empresa
        empleado = Employee.query.get(ausencia.employee_id)
        if empleado.company_id != company_id:
            flash('La ausencia no pertenece a un empleado de esta empresa.', 'danger')
            return redirect(url_for('turnos.listar_ausencias', company_id=company_id))
        
        ausencia.aprobado = form.aprobado.data
        ausencia.aprobado_por_id = current_user.id
        db.session.commit()
        
        estado = "aprobada" if form.aprobado.data else "rechazada"
        flash(f'Ausencia {estado} correctamente.', 'success')
    else:
        flash('No se pudo procesar la solicitud de ausencia.', 'danger')
    
    return redirect(url_for('turnos.listar_ausencias', company_id=company_id))

@turnos_bp.route('/ausencias/<int:company_id>/eliminar/<int:ausencia_id>', methods=['POST'])
@login_required
def eliminar_ausencia(company_id, ausencia_id):
    """Elimina una ausencia."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    ausencia = Ausencia.query.get_or_404(ausencia_id)
    
    # Verificar que el empleado pertenece a la empresa
    empleado = Employee.query.get(ausencia.employee_id)
    if empleado.company_id != company_id:
        flash('La ausencia no pertenece a un empleado de esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_ausencias', company_id=company_id))
    
    db.session.delete(ausencia)
    db.session.commit()
    
    flash('Ausencia eliminada correctamente.', 'success')
    return redirect(url_for('turnos.listar_ausencias', company_id=company_id))

# Requisitos de personal
@turnos_bp.route('/requisitos/<int:company_id>')
@login_required
def listar_requisitos(company_id):
    """Lista los requisitos de personal de una empresa."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    requisitos = RequisitoPersonal.query.filter_by(company_id=company_id, is_active=True).order_by(
        RequisitoPersonal.dia_semana, RequisitoPersonal.hora_inicio
    ).all()
    
    return render_template('turnos/listar_requisitos.html',
                          title=f'Requisitos de Personal - {company.name}',
                          company=company,
                          requisitos=requisitos,
                          dias_semana=dias_semana)

@turnos_bp.route('/requisitos/<int:company_id>/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_requisito(company_id):
    """Crea un nuevo requisito de personal."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = RequisitoPersonalForm()
    
    if form.validate_on_submit():
        requisito = RequisitoPersonal(
            dia_semana=form.dia_semana.data,
            hora_inicio=form.hora_inicio.data,
            hora_fin=form.hora_fin.data,
            num_empleados=form.num_empleados.data,
            notas=form.notas.data,
            company_id=company_id
        )
        
        db.session.add(requisito)
        db.session.commit()
        
        flash('Requisito de personal creado correctamente.', 'success')
        return redirect(url_for('turnos.listar_requisitos', company_id=company_id))
    
    form.company_id.data = company_id
    
    return render_template('turnos/form_requisito.html',
                          title=f'Nuevo Requisito de Personal - {company.name}',
                          company=company,
                          form=form)

@turnos_bp.route('/requisitos/<int:company_id>/editar/<int:requisito_id>', methods=['GET', 'POST'])
@login_required
def editar_requisito(company_id, requisito_id):
    """Edita un requisito de personal existente."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    requisito = RequisitoPersonal.query.get_or_404(requisito_id)
    
    if requisito.company_id != company_id:
        flash('El requisito no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_requisitos', company_id=company_id))
    
    form = RequisitoPersonalForm(obj=requisito)
    
    if form.validate_on_submit():
        form.populate_obj(requisito)
        db.session.commit()
        
        flash('Requisito de personal actualizado correctamente.', 'success')
        return redirect(url_for('turnos.listar_requisitos', company_id=company_id))
    
    form.company_id.data = company_id
    
    return render_template('turnos/form_requisito.html',
                          title=f'Editar Requisito de Personal - {company.name}',
                          company=company,
                          form=form,
                          requisito=requisito)

@turnos_bp.route('/requisitos/<int:company_id>/eliminar/<int:requisito_id>', methods=['POST'])
@login_required
def eliminar_requisito(company_id, requisito_id):
    """Elimina un requisito de personal (marcar como inactivo)."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    requisito = RequisitoPersonal.query.get_or_404(requisito_id)
    
    if requisito.company_id != company_id:
        flash('El requisito no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_requisitos', company_id=company_id))
    
    requisito.is_active = False
    db.session.commit()
    
    flash('Requisito de personal eliminado correctamente.', 'success')
    return redirect(url_for('turnos.listar_requisitos', company_id=company_id))

# Plantillas de horarios
@turnos_bp.route('/plantillas/<int:company_id>')
@login_required
def listar_plantillas(company_id):
    """Lista las plantillas de horarios de una empresa."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    plantillas = PlantillaHorario.query.filter_by(company_id=company_id, is_active=True).order_by(PlantillaHorario.nombre).all()
    
    return render_template('turnos/listar_plantillas.html',
                          title=f'Plantillas de Horarios - {company.name}',
                          company=company,
                          plantillas=plantillas)

@turnos_bp.route('/plantillas/<int:company_id>/nueva', methods=['GET', 'POST'])
@login_required
def nueva_plantilla(company_id):
    """Crea una nueva plantilla de horarios."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = PlantillaHorarioForm()
    
    if form.validate_on_submit():
        plantilla = PlantillaHorario(
            nombre=form.nombre.data,
            descripcion=form.descripcion.data,
            company_id=company_id
        )
        
        db.session.add(plantilla)
        db.session.commit()
        
        flash(f'Plantilla "{plantilla.nombre}" creada correctamente.', 'success')
        return redirect(url_for('turnos.editar_plantilla', company_id=company_id, plantilla_id=plantilla.id))
    
    form.company_id.data = company_id
    
    return render_template('turnos/form_plantilla.html',
                          title=f'Nueva Plantilla de Horarios - {company.name}',
                          company=company,
                          form=form)

@turnos_bp.route('/plantillas/<int:company_id>/ver/<int:plantilla_id>')
@login_required
def ver_plantilla(company_id, plantilla_id):
    """Muestra los detalles de una plantilla de horarios."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    plantilla = PlantillaHorario.query.get_or_404(plantilla_id)
    
    if plantilla.company_id != company_id:
        flash('La plantilla no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    # Organizar detalles por día de la semana
    detalles_por_dia = {dia: [] for dia in range(7)}
    for detalle in plantilla.detalles:
        detalles_por_dia[detalle.dia_semana].append(detalle)
    
    # Obtener asignaciones de esta plantilla
    asignaciones = AsignacionPlantilla.query.filter_by(plantilla_id=plantilla_id, is_active=True).all()
    
    return render_template('turnos/ver_plantilla.html',
                          title=f'Plantilla: {plantilla.nombre} - {company.name}',
                          company=company,
                          plantilla=plantilla,
                          detalles_por_dia=detalles_por_dia,
                          dias_semana=dias_semana,
                          asignaciones=asignaciones)

@turnos_bp.route('/plantillas/<int:company_id>/editar/<int:plantilla_id>', methods=['GET', 'POST'])
@login_required
def editar_plantilla(company_id, plantilla_id):
    """Edita una plantilla de horarios existente."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    plantilla = PlantillaHorario.query.get_or_404(plantilla_id)
    
    if plantilla.company_id != company_id:
        flash('La plantilla no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    # Formulario para editar los datos básicos de la plantilla
    form = PlantillaHorarioForm(obj=plantilla)
    
    if form.validate_on_submit():
        form.populate_obj(plantilla)
        db.session.commit()
        
        flash(f'Plantilla "{plantilla.nombre}" actualizada correctamente.', 'success')
        return redirect(url_for('turnos.editar_plantilla', company_id=company_id, plantilla_id=plantilla.id))
    
    # Formulario para agregar un nuevo detalle a la plantilla
    detalle_form = DetallePlantillaForm()
    turnos = get_turnos_empresa(company_id)
    detalle_form.turno_id.choices = [(t.id, t.nombre) for t in turnos]
    detalle_form.plantilla_id.data = plantilla_id
    
    # Organizar detalles por día de la semana
    detalles_por_dia = {dia: [] for dia in range(7)}
    for detalle in plantilla.detalles:
        detalles_por_dia[detalle.dia_semana].append(detalle)
    
    return render_template('turnos/editar_plantilla.html',
                          title=f'Editar Plantilla: {plantilla.nombre} - {company.name}',
                          company=company,
                          plantilla=plantilla,
                          form=form,
                          detalle_form=detalle_form,
                          detalles_por_dia=detalles_por_dia,
                          dias_semana=dias_semana)

@turnos_bp.route('/plantillas/<int:company_id>/agregar_detalle/<int:plantilla_id>', methods=['POST'])
@login_required
def agregar_detalle_plantilla(company_id, plantilla_id):
    """Agrega un detalle a una plantilla de horarios."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    plantilla = PlantillaHorario.query.get_or_404(plantilla_id)
    
    if plantilla.company_id != company_id:
        flash('La plantilla no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    form = DetallePlantillaForm()
    turnos = get_turnos_empresa(company_id)
    form.turno_id.choices = [(t.id, t.nombre) for t in turnos]
    
    if form.validate_on_submit():
        # Verificar si ya existe un detalle para este día y turno
        existe = DetallePlantilla.query.filter_by(
            plantilla_id=plantilla_id,
            dia_semana=form.dia_semana.data,
            turno_id=form.turno_id.data
        ).first()
        
        if existe:
            flash('Ya existe este turno para este día en la plantilla.', 'warning')
            return redirect(url_for('turnos.editar_plantilla', company_id=company_id, plantilla_id=plantilla_id))
        
        detalle = DetallePlantilla(
            dia_semana=form.dia_semana.data,
            turno_id=form.turno_id.data,
            plantilla_id=plantilla_id
        )
        
        db.session.add(detalle)
        db.session.commit()
        
        flash('Turno agregado a la plantilla correctamente.', 'success')
    else:
        flash('Error al agregar el turno a la plantilla.', 'danger')
    
    return redirect(url_for('turnos.editar_plantilla', company_id=company_id, plantilla_id=plantilla_id))

@turnos_bp.route('/plantillas/<int:company_id>/eliminar_detalle/<int:detalle_id>', methods=['POST'])
@login_required
def eliminar_detalle_plantilla(company_id, detalle_id):
    """Elimina un detalle de una plantilla de horarios."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    detalle = DetallePlantilla.query.get_or_404(detalle_id)
    plantilla_id = detalle.plantilla_id
    plantilla = PlantillaHorario.query.get_or_404(plantilla_id)
    
    if plantilla.company_id != company_id:
        flash('La plantilla no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    db.session.delete(detalle)
    db.session.commit()
    
    flash('Turno eliminado de la plantilla correctamente.', 'success')
    return redirect(url_for('turnos.editar_plantilla', company_id=company_id, plantilla_id=plantilla_id))

@turnos_bp.route('/plantillas/<int:company_id>/eliminar/<int:plantilla_id>', methods=['POST'])
@login_required
def eliminar_plantilla(company_id, plantilla_id):
    """Elimina una plantilla de horarios (marcar como inactiva)."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    plantilla = PlantillaHorario.query.get_or_404(plantilla_id)
    
    if plantilla.company_id != company_id:
        flash('La plantilla no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    # Verificar si hay asignaciones activas de esta plantilla
    asignaciones_count = AsignacionPlantilla.query.filter_by(plantilla_id=plantilla_id, is_active=True).count()
    if asignaciones_count > 0:
        flash(f'No se puede eliminar la plantilla "{plantilla.nombre}" porque tiene {asignaciones_count} asignaciones activas.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    plantilla.is_active = False
    db.session.commit()
    
    flash(f'Plantilla "{plantilla.nombre}" eliminada correctamente.', 'success')
    return redirect(url_for('turnos.listar_plantillas', company_id=company_id))

# Asignación de plantillas
@turnos_bp.route('/asignar_plantilla/<int:company_id>', methods=['GET', 'POST'])
@login_required
def asignar_plantilla(company_id):
    """Asigna una plantilla de horarios a un empleado."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = AsignacionPlantillaForm()
    
    # Cargar empleados y plantillas para los select
    empleados = get_empleados_empresa(company_id)
    plantillas = get_plantillas_empresa(company_id)
    
    form.employee_id.choices = [(e.id, f"{e.nombre} {e.apellidos}") for e in empleados]
    form.plantilla_id.choices = [(p.id, p.nombre) for p in plantillas]
    
    if form.validate_on_submit():
        asignacion = AsignacionPlantilla(
            plantilla_id=form.plantilla_id.data,
            employee_id=form.employee_id.data,
            fecha_inicio=form.fecha_inicio.data,
            fecha_fin=form.fecha_fin.data
        )
        
        db.session.add(asignacion)
        db.session.commit()
        
        empleado = Employee.query.get(form.employee_id.data)
        plantilla = PlantillaHorario.query.get(form.plantilla_id.data)
        flash(f'Plantilla "{plantilla.nombre}" asignada correctamente a {empleado.nombre} {empleado.apellidos}.', 'success')
        
        # Preguntar si desea generar los horarios ahora
        return redirect(url_for('turnos.confirmar_generar_horarios', 
                              company_id=company_id, 
                              asignacion_id=asignacion.id))
    
    # Si viene con un empleado o plantilla preseleccionados
    employee_id = request.args.get('employee_id', type=int)
    if employee_id:
        form.employee_id.data = employee_id
    
    plantilla_id = request.args.get('plantilla_id', type=int)
    if plantilla_id:
        form.plantilla_id.data = plantilla_id
    
    # Valores por defecto
    if request.method == 'GET':
        form.fecha_inicio.data = date.today()
    
    return render_template('turnos/form_asignar_plantilla.html',
                          title=f'Asignar Plantilla de Horarios - {company.name}',
                          company=company,
                          form=form)

@turnos_bp.route('/confirmar_generar_horarios/<int:company_id>/<int:asignacion_id>')
@login_required
def confirmar_generar_horarios(company_id, asignacion_id):
    """Confirmación para generar horarios a partir de una asignación de plantilla."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    asignacion = AsignacionPlantilla.query.get_or_404(asignacion_id)
    empleado = Employee.query.get(asignacion.employee_id)
    plantilla = PlantillaHorario.query.get(asignacion.plantilla_id)
    
    if empleado.company_id != company_id:
        flash('El empleado no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    return render_template('turnos/confirmar_generar_horarios.html',
                          title=f'Generar Horarios - {company.name}',
                          company=company,
                          asignacion=asignacion,
                          empleado=empleado,
                          plantilla=plantilla)

@turnos_bp.route('/generar_horarios/<int:company_id>/<int:asignacion_id>', methods=['POST'])
@login_required
def generar_horarios(company_id, asignacion_id):
    """Genera horarios a partir de una asignación de plantilla."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    asignacion = AsignacionPlantilla.query.get_or_404(asignacion_id)
    empleado = Employee.query.get(asignacion.employee_id)
    plantilla = PlantillaHorario.query.get(asignacion.plantilla_id)
    
    if empleado.company_id != company_id:
        flash('El empleado no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    # Obtener el rango de fechas desde el formulario
    fecha_inicio_str = request.form.get('fecha_inicio')
    fecha_fin_str = request.form.get('fecha_fin')
    
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else None
    except ValueError:
        flash('Las fechas proporcionadas no son válidas.', 'danger')
        return redirect(url_for('turnos.confirmar_generar_horarios', company_id=company_id, asignacion_id=asignacion_id))
    
    if fecha_fin and fecha_fin < fecha_inicio:
        flash('La fecha de fin debe ser posterior a la fecha de inicio.', 'danger')
        return redirect(url_for('turnos.confirmar_generar_horarios', company_id=company_id, asignacion_id=asignacion_id))
    
    # Si no se proporciona fecha de fin, usar 4 semanas como máximo
    if not fecha_fin:
        fecha_fin = fecha_inicio + timedelta(days=28)
    
    # Obtener detalles de la plantilla
    detalles_por_dia = {}
    for detalle in plantilla.detalles:
        if detalle.dia_semana not in detalles_por_dia:
            detalles_por_dia[detalle.dia_semana] = []
        detalles_por_dia[detalle.dia_semana].append(detalle)
    
    # Generar horarios
    horarios_creados = 0
    horarios_omitidos = 0
    
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        dia_semana = fecha_actual.weekday()
        
        # Si hay turnos definidos para este día de la semana
        if dia_semana in detalles_por_dia:
            for detalle in detalles_por_dia[dia_semana]:
                # Verificar si ya existe una asignación para esta fecha
                existe = Horario.query.filter_by(
                    employee_id=empleado.id,
                    fecha=fecha_actual,
                    is_active=True
                ).first()
                
                if not existe:
                    horario = Horario(
                        fecha=fecha_actual,
                        turno_id=detalle.turno_id,
                        employee_id=empleado.id,
                        notas=f"Generado automáticamente desde plantilla '{plantilla.nombre}'"
                    )
                    
                    db.session.add(horario)
                    horarios_creados += 1
                    
                    # Registrar el cambio en el historial
                    turno = Turno.query.get(detalle.turno_id)
                    registrar_cambio_horario(
                        horario,
                        "creacion",
                        f"Asignación automática de turno '{turno.nombre}' a {empleado.nombre} {empleado.apellidos} para el {fecha_actual.strftime('%d/%m/%Y')} desde plantilla '{plantilla.nombre}'"
                    )
                else:
                    horarios_omitidos += 1
        
        fecha_actual += timedelta(days=1)
    
    db.session.commit()
    
    mensaje = f'Se han generado {horarios_creados} horarios correctamente a partir de la plantilla.'
    if horarios_omitidos > 0:
        mensaje += f' Se omitieron {horarios_omitidos} asignaciones porque ya existían.'
    
    flash(mensaje, 'success')
    
    # Redirigir al calendario en la semana de la fecha inicial
    inicio_semana = fecha_inicio - timedelta(days=fecha_inicio.weekday())
    return redirect(url_for('turnos.calendario', company_id=company_id, semana=inicio_semana.strftime('%Y-%m-%d')))

@turnos_bp.route('/desasignar_plantilla/<int:company_id>/<int:asignacion_id>', methods=['POST'])
@login_required
def desasignar_plantilla(company_id, asignacion_id):
    """Desasigna una plantilla de un empleado (marcar como inactiva)."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    asignacion = AsignacionPlantilla.query.get_or_404(asignacion_id)
    empleado = Employee.query.get(asignacion.employee_id)
    plantilla = PlantillaHorario.query.get(asignacion.plantilla_id)
    
    if empleado.company_id != company_id:
        flash('El empleado no pertenece a esta empresa.', 'danger')
        return redirect(url_for('turnos.listar_plantillas', company_id=company_id))
    
    asignacion.is_active = False
    db.session.commit()
    
    flash(f'Plantilla "{plantilla.nombre}" desasignada correctamente de {empleado.nombre} {empleado.apellidos}.', 'success')
    return redirect(url_for('turnos.ver_plantilla', company_id=company_id, plantilla_id=plantilla.id))

# Informes y exportación
@turnos_bp.route('/informes/<int:company_id>')
@login_required
def informes(company_id):
    """Página de informes del módulo de turnos."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    
    return render_template('turnos/informes.html',
                          title=f'Informes de Turnos - {company.name}',
                          company=company)

@turnos_bp.route('/informe_horarios/<int:company_id>', methods=['GET', 'POST'])
@login_required
def informe_horarios(company_id):
    """Genera un informe de horarios."""
    if not puede_gestionar_empresa(company_id):
        flash('No tiene permiso para gestionar esta empresa.', 'danger')
        return redirect(url_for('turnos.select_company'))
    
    company = Company.query.get_or_404(company_id)
    form = FiltroHorarioForm()
    
    # Cargar empleados y turnos para los select
    empleados = get_empleados_empresa(company_id)
    turnos = get_turnos_empresa(company_id)
    
    form.employee_id.choices = [(-1, 'Todos los empleados')] + [(e.id, f"{e.nombre} {e.apellidos}") for e in empleados]
    form.turno_id.choices = [(-1, 'Todos los turnos')] + [(t.id, t.nombre) for t in turnos]
    
    if form.validate_on_submit() or request.method == 'GET' and request.args.get('fecha_inicio'):
        # Si es POST o GET con parámetros, procesar el informe
        if request.method == 'GET' and request.args.get('fecha_inicio'):
            # Obtener filtros de la URL
            try:
                fecha_inicio = datetime.strptime(request.args.get('fecha_inicio'), '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(request.args.get('fecha_fin'), '%Y-%m-%d').date() if request.args.get('fecha_fin') else None
                employee_id = int(request.args.get('employee_id', -1))
                turno_id = int(request.args.get('turno_id', -1))
            except (ValueError, TypeError):
                flash('Parámetros de filtro inválidos.', 'danger')
                return redirect(url_for('turnos.informe_horarios', company_id=company_id))
        else:
            # Obtener filtros del formulario
            fecha_inicio = form.fecha_inicio.data
            fecha_fin = form.fecha_fin.data
            employee_id = form.employee_id.data
            turno_id = form.turno_id.data
        
        # Si no se proporciona fecha de fin, usar fecha de inicio + 30 días
        if not fecha_fin:
            fecha_fin = fecha_inicio + timedelta(days=30)
        
        # Construir la consulta base
        query = Horario.query.join(Employee).filter(
            Employee.company_id == company_id,
            Horario.fecha >= fecha_inicio,
            Horario.fecha <= fecha_fin,
            Horario.is_active == True
        )
        
        # Aplicar filtros adicionales si se seleccionaron
        if employee_id > 0:
            query = query.filter(Horario.employee_id == employee_id)
        
        if turno_id > 0:
            query = query.filter(Horario.turno_id == turno_id)
        
        # Ordenar por fecha y empleado
        horarios = query.order_by(Horario.fecha, Horario.employee_id).all()
        
        # Agrupar horarios por empleado
        horarios_por_empleado = {}
        for horario in horarios:
            if horario.employee_id not in horarios_por_empleado:
                horarios_por_empleado[horario.employee_id] = {
                    'empleado': horario.employee,
                    'horarios': [],
                    'horas_totales': 0,
                    'dias_trabajados': 0
                }
            
            horarios_por_empleado[horario.employee_id]['horarios'].append(horario)
            horarios_por_empleado[horario.employee_id]['horas_totales'] += horario.turno.horas_efectivas
            horarios_por_empleado[horario.employee_id]['dias_trabajados'] += 1
        
        # Exportar si se solicita
        formato = request.args.get('formato') or request.form.get('formato')
        if formato:
            return exportar_informe_horarios(
                company, 
                horarios, 
                horarios_por_empleado, 
                fecha_inicio, 
                fecha_fin, 
                formato
            )
        
        # Valores para el formulario al redireccionar por GET
        if request.method == 'GET' and request.args.get('fecha_inicio'):
            form.fecha_inicio.data = fecha_inicio
            form.fecha_fin.data = fecha_fin
            form.employee_id.data = employee_id
            form.turno_id.data = turno_id
        
        return render_template('turnos/informe_horarios.html',
                              title=f'Informe de Horarios - {company.name}',
                              company=company,
                              form=form,
                              horarios=horarios,
                              horarios_por_empleado=horarios_por_empleado,
                              fecha_inicio=fecha_inicio,
                              fecha_fin=fecha_fin,
                              hay_resultados=len(horarios) > 0)
    
    # Valores por defecto para el formulario
    if request.method == 'GET' and not request.args.get('fecha_inicio'):
        form.fecha_inicio.data = date.today().replace(day=1)  # Primer día del mes actual
        form.fecha_fin.data = date.today().replace(day=calendar.monthrange(date.today().year, date.today().month)[1])  # Último día del mes
        form.employee_id.data = -1  # Todos los empleados
        form.turno_id.data = -1  # Todos los turnos
    
    return render_template('turnos/informe_horarios.html',
                          title=f'Informe de Horarios - {company.name}',
                          company=company,
                          form=form,
                          hay_resultados=False)

def exportar_informe_horarios(company, horarios, horarios_por_empleado, fecha_inicio, fecha_fin, formato):
    """Exporta un informe de horarios en el formato especificado."""
    if formato == 'csv':
        # Crear archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Encabezados
        writer.writerow(['Empleado', 'Fecha', 'Día', 'Turno', 'Hora Inicio', 'Hora Fin', 'Horas', 'Notas'])
        
        # Datos
        for horario in horarios:
            dia_semana = dias_semana[horario.fecha.weekday()]
            writer.writerow([
                f"{horario.employee.nombre} {horario.employee.apellidos}",
                horario.fecha.strftime('%d/%m/%Y'),
                dia_semana,
                horario.turno.nombre,
                horario.turno.hora_inicio.strftime('%H:%M'),
                horario.turno.hora_fin.strftime('%H:%M'),
                f"{horario.turno.horas_efectivas:.2f}",
                horario.notas or ''
            ])
        
        # Configurar respuesta
        output.seek(0)
        filename = f"horarios_{company.name}_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.csv"
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
    
    elif formato == 'excel':
        # Crear libro Excel en memoria
        wb = Workbook()
        ws = wb.active
        ws.title = "Horarios"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ['Empleado', 'Fecha', 'Día', 'Turno', 'Hora Inicio', 'Hora Fin', 'Horas', 'Notas']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        row = 2
        for horario in horarios:
            dia_semana = dias_semana[horario.fecha.weekday()]
            ws.cell(row=row, column=1, value=f"{horario.employee.nombre} {horario.employee.apellidos}")
            ws.cell(row=row, column=2, value=horario.fecha.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=dia_semana)
            ws.cell(row=row, column=4, value=horario.turno.nombre)
            ws.cell(row=row, column=5, value=horario.turno.hora_inicio.strftime('%H:%M'))
            ws.cell(row=row, column=6, value=horario.turno.hora_fin.strftime('%H:%M'))
            ws.cell(row=row, column=7, value=round(horario.turno.horas_efectivas, 2))
            ws.cell(row=row, column=8, value=horario.notas or '')
            row += 1
        
        # Autajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        
        # Agregar hoja de resumen
        ws_resumen = wb.create_sheet(title="Resumen")
        
        # Encabezados resumen
        headers_resumen = ['Empleado', 'Días Trabajados', 'Horas Totales']
        for col, header in enumerate(headers_resumen, start=1):
            cell = ws_resumen.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos resumen
        row = 2
        for datos in horarios_por_empleado.values():
            ws_resumen.cell(row=row, column=1, value=f"{datos['empleado'].nombre} {datos['empleado'].apellidos}")
            ws_resumen.cell(row=row, column=2, value=datos['dias_trabajados'])
            ws_resumen.cell(row=row, column=3, value=round(datos['horas_totales'], 2))
            row += 1
        
        # Autajustar columnas resumen
        for col in ws_resumen.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws_resumen.column_dimensions[column].width = adjusted_width
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Configurar respuesta
        filename = f"horarios_{company.name}_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    elif formato == 'pdf':
        from fpdf import FPDF
        
        class PDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 12)
                self.cell(0, 10, f'Informe de Horarios - {company.name}', 0, 1, 'C')
                self.cell(0, 10, f'Período: {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}', 0, 1, 'C')
                self.ln(5)
            
            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')
        
        # Crear PDF
        pdf = PDF()
        pdf.add_page()
        
        # Resumen por empleado
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'Resumen por Empleado', 0, 1)
        pdf.ln(2)
        
        # Tabla de resumen
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(90, 7, 'Empleado', 1, 0, 'C')
        pdf.cell(30, 7, 'Días', 1, 0, 'C')
        pdf.cell(30, 7, 'Horas', 1, 1, 'C')
        
        pdf.set_font('Arial', '', 10)
        for datos in horarios_por_empleado.values():
            pdf.cell(90, 7, f"{datos['empleado'].nombre} {datos['empleado'].apellidos}", 1, 0)
            pdf.cell(30, 7, str(datos['dias_trabajados']), 1, 0, 'C')
            pdf.cell(30, 7, f"{datos['horas_totales']:.2f}", 1, 1, 'C')
        
        pdf.ln(10)
        
        # Detalle de horarios
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'Detalle de Horarios', 0, 1)
        pdf.ln(2)
        
        # Agrupar por empleado para el detalle
        for employee_id, datos in horarios_por_empleado.items():
            pdf.set_font('Arial', 'B', 11)
            pdf.cell(0, 7, f"{datos['empleado'].nombre} {datos['empleado'].apellidos}", 0, 1)
            pdf.ln(2)
            
            # Tabla de horarios por empleado
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(25, 7, 'Fecha', 1, 0, 'C')
            pdf.cell(20, 7, 'Día', 1, 0, 'C')
            pdf.cell(30, 7, 'Turno', 1, 0, 'C')
            pdf.cell(20, 7, 'Inicio', 1, 0, 'C')
            pdf.cell(20, 7, 'Fin', 1, 0, 'C')
            pdf.cell(15, 7, 'Horas', 1, 0, 'C')
            pdf.cell(60, 7, 'Notas', 1, 1, 'C')
            
            pdf.set_font('Arial', '', 8)
            for horario in datos['horarios']:
                dia_semana = dias_semana[horario.fecha.weekday()]
                
                # Comprobar si hay suficiente espacio en la página actual
                if pdf.get_y() + 7 > pdf.page_break_trigger:
                    pdf.add_page()
                    
                    # Repetir encabezados de la tabla
                    pdf.set_font('Arial', 'B', 11)
                    pdf.cell(0, 7, f"{datos['empleado'].nombre} {datos['empleado'].apellidos} (cont.)", 0, 1)
                    pdf.ln(2)
                    
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(25, 7, 'Fecha', 1, 0, 'C')
                    pdf.cell(20, 7, 'Día', 1, 0, 'C')
                    pdf.cell(30, 7, 'Turno', 1, 0, 'C')
                    pdf.cell(20, 7, 'Inicio', 1, 0, 'C')
                    pdf.cell(20, 7, 'Fin', 1, 0, 'C')
                    pdf.cell(15, 7, 'Horas', 1, 0, 'C')
                    pdf.cell(60, 7, 'Notas', 1, 1, 'C')
                    
                    pdf.set_font('Arial', '', 8)
                
                pdf.cell(25, 7, horario.fecha.strftime('%d/%m/%Y'), 1, 0)
                pdf.cell(20, 7, dia_semana, 1, 0)
                pdf.cell(30, 7, horario.turno.nombre, 1, 0)
                pdf.cell(20, 7, horario.turno.hora_inicio.strftime('%H:%M'), 1, 0, 'C')
                pdf.cell(20, 7, horario.turno.hora_fin.strftime('%H:%M'), 1, 0, 'C')
                pdf.cell(15, 7, f"{horario.turno.horas_efectivas:.2f}", 1, 0, 'C')
                
                # Gestionar texto largo en las notas
                notas = horario.notas or ''
                if len(notas) > 28:
                    notas = notas[:25] + '...'
                pdf.cell(60, 7, notas, 1, 1)
            
            pdf.ln(5)
        
        # Guardar en memoria
        output = io.BytesIO()
        output.write(pdf.output(dest='S').encode('latin1'))
        output.seek(0)
        
        # Configurar respuesta
        filename = f"horarios_{company.name}_{fecha_inicio.strftime('%Y%m%d')}_{fecha_fin.strftime('%Y%m%d')}.pdf"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
    
    # Si el formato no es válido, redirigir al informe
    flash('Formato de exportación no válido.', 'danger')
    return redirect(url_for('turnos.informe_horarios', 
                        company_id=company.id,
                        fecha_inicio=fecha_inicio.strftime('%Y-%m-%d'),
                        fecha_fin=fecha_fin.strftime('%Y-%m-%d')))

# API para interacción en el frontend
@turnos_bp.route('/api/horarios/<int:company_id>/<fecha>')
@login_required
def api_horarios_fecha(company_id, fecha):
    """API para obtener horarios de una fecha específica."""
    if not puede_gestionar_empresa(company_id):
        return jsonify({'error': 'No tiene permiso para acceder a estos datos'}), 403
    
    try:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Formato de fecha inválido'}), 400
    
    # Obtener horarios para la fecha
    horarios = Horario.query.join(Employee).filter(
        Employee.company_id == company_id,
        Horario.fecha == fecha_obj,
        Horario.is_active == True
    ).all()
    
    # Formatear resultados
    resultados = []
    for horario in horarios:
        resultados.append({
            'id': horario.id,
            'employee_id': horario.employee_id,
            'employee_nombre': f"{horario.employee.nombre} {horario.employee.apellidos}",
            'turno_id': horario.turno_id,
            'turno_nombre': horario.turno.nombre,
            'turno_color': horario.turno.color,
            'hora_inicio': horario.turno.hora_inicio.strftime('%H:%M'),
            'hora_fin': horario.turno.hora_fin.strftime('%H:%M'),
            'notas': horario.notas
        })
    
    return jsonify(resultados)

def init_app(app):
    """Inicializa la aplicación con este blueprint."""
    app.register_blueprint(turnos_bp)
    
    # Agrega el módulo de turnos a la navegación
    @app.context_processor
    def inject_turnos_nav():
        return {
            'show_turnos_nav': True
        }